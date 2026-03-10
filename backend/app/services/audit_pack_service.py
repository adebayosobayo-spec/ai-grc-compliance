"""
Audit Pack Export Service.

Builds a ZIP archive containing:
  - registers/risk_register.xlsx
  - registers/asset_register.xlsx
  - registers/supplier_register.xlsx
  - registers/data_processing_register.xlsx
  - registers/ai_system_register.xlsx  (ISO 42001 only)
  - registers/statement_of_applicability.xlsx
  - evidence/evidence_register.xlsx
  - README.txt  (cover sheet with org name, date, framework)

The ZIP is returned as bytes for streaming to the client.
"""
import io
import zipfile
from datetime import datetime
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session

from app.services import register_service


# ── Style helpers ──────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_CELL_FONT = Font(name="Calibri", size=10)
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ALT_FILL = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")


def _write_sheet(ws, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    header_row = ws[1]
    for cell in header_row:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _BORDER

    for ri, row in enumerate(rows):
        ws.append(row)
        fill = _ALT_FILL if ri % 2 == 0 else None
        for ci, cell in enumerate(ws[ri + 2]):
            cell.font = _CELL_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


# ── Individual sheet builders ──────────────────────────────────────────────────

def _build_risk_sheet(ws, risks) -> None:
    headers = ["Risk ID", "Description", "Asset", "Threat", "Vulnerability",
               "Likelihood", "Impact", "Risk Level", "Treatment", "Controls", "Owner", "Status", "Notes"]
    rows = [
        [r.risk_id, r.risk_description, r.asset, r.threat, r.vulnerability,
         r.likelihood, r.impact, r.risk_level, r.treatment, r.control_refs,
         r.owner, r.status, r.notes]
        for r in risks
    ]
    _write_sheet(ws, headers, rows)


def _build_asset_sheet(ws, assets) -> None:
    headers = ["Asset ID", "Name", "Type", "Classification", "Owner", "Location", "Value", "Controls", "Notes"]
    rows = [
        [a.asset_id, a.asset_name, a.asset_type, a.classification,
         a.owner, a.location, a.value, a.control_refs, a.notes]
        for a in assets
    ]
    _write_sheet(ws, headers, rows)


def _build_supplier_sheet(ws, suppliers) -> None:
    headers = ["Supplier ID", "Name", "Service", "Data Access", "Risk Rating",
               "Contract Status", "Review Date", "Contact", "Notes"]
    rows = [
        [s.supplier_id, s.supplier_name, s.service_provided, s.data_access,
         s.risk_rating, s.contract_status, s.review_date, s.contact, s.notes]
        for s in suppliers
    ]
    _write_sheet(ws, headers, rows)


def _build_data_processing_sheet(ws, records) -> None:
    headers = ["Record ID", "Processing Activity", "Purpose", "Legal Basis",
               "Data Categories", "Data Subjects", "Retention", "Third-Party Transfers", "Security Measures", "Notes"]
    rows = [
        [r.record_id, r.processing_activity, r.purpose, r.legal_basis,
         r.data_categories, r.data_subjects, r.retention_period,
         r.third_party_transfers, r.security_measures, r.notes]
        for r in records
    ]
    _write_sheet(ws, headers, rows)


def _build_ai_system_sheet(ws, systems) -> None:
    headers = ["System ID", "Name", "Purpose", "Risk Classification", "AI Type",
               "Training Data", "Output Type", "Human Oversight", "Vendor", "Status", "Notes"]
    rows = [
        [s.system_id, s.system_name, s.purpose, s.risk_classification, s.ai_type,
         s.training_data, s.output_type, s.human_oversight, s.vendor, s.status, s.notes]
        for s in systems
    ]
    _write_sheet(ws, headers, rows)


def _build_soa_sheet(ws, controls) -> None:
    headers = ["Control ID", "Control Name", "Applicable", "Justification",
               "Implementation Status", "Evidence Summary", "Owner", "Target Date", "Notes"]
    rows = [
        [c.control_id, c.control_name, "Yes" if c.applicable else "No",
         c.justification, c.implementation_status, c.evidence_summary,
         c.owner, c.target_date, c.notes]
        for c in controls
    ]
    _write_sheet(ws, headers, rows)


def _build_evidence_sheet(ws, evidence) -> None:
    headers = ["Evidence ID", "Control ID", "Type", "Title", "Description",
               "File Reference", "Status", "Reviewer", "Notes"]
    rows = [
        [e.evidence_id, e.control_id, e.evidence_type, e.title, e.description,
         e.file_reference, e.status, e.reviewer, e.notes]
        for e in evidence
    ]
    _write_sheet(ws, headers, rows)


# ── Excel workbook builder ─────────────────────────────────────────────────────

def _build_excel(sheet_name: str, build_fn, items) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    build_fn(ws, items)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── README cover sheet ─────────────────────────────────────────────────────────

def _build_readme(org_name: str, framework: str, session_id: str) -> str:
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    _FW_LABELS = {
        "ISO_27001": "ISO/IEC 27001:2022",
        "ISO_42001": "ISO/IEC 42001:2023",
        "GDPR": "EU GDPR",
        "UK_GDPR": "UK GDPR",
        "NDPR": "Nigeria NDPR",
        "POPIA": "South Africa POPIA",
        "LGPD": "Brazil LGPD",
        "CCPA": "California CCPA/CPRA",
        "PDPA": "Singapore/Thailand PDPA",
    }
    fw_label = _FW_LABELS.get(framework, framework)
    return f"""COMPLAI — Audit Pack
{'=' * 60}

Organisation:  {org_name}
Framework:     {fw_label}
Session ID:    {session_id}
Generated:     {now}

CONTENTS
--------
registers/
  risk_register.xlsx               — Risk Register (ISO {fw_label})
  asset_register.xlsx              — Asset / Information Asset Register
  supplier_register.xlsx           — Third-Party / Supplier Register
  data_processing_register.xlsx    — Data Processing Activities Register
  {'ai_system_register.xlsx          — AI System Inventory (ISO 42001 only)' if framework == 'ISO_42001' else ''}
  statement_of_applicability.xlsx  — Statement of Applicability (SoA)
evidence/
  evidence_register.xlsx           — Evidence Artefacts Linked to Controls

NOTES
-----
- All registers were seeded by the COMPLAI onboarding wizard and may
  require review and enrichment before submission to an auditor.
- This pack does not constitute a complete audit submission.
  Consult a qualified ISO auditor for certification readiness.
- Generated by COMPLAI v1.0 — Compliance Intelligence Platform.
"""


# ── Main export function ───────────────────────────────────────────────────────

def build_audit_pack(db: Session, session_id: str, org_name: str, framework: str) -> bytes:
    """
    Build a ZIP archive of all register Excel files for the given session.
    Returns the ZIP as bytes.
    """
    risks = register_service.list_risks(db, session_id)
    assets = register_service.list_assets(db, session_id)
    suppliers = register_service.list_suppliers(db, session_id)
    dp_records = register_service.list_data_processing(db, session_id)
    ai_systems = register_service.list_ai_systems(db, session_id)
    controls = register_service.list_controls(db, session_id)
    evidence = register_service.list_evidence(db, session_id)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("README.txt", _build_readme(org_name, framework, session_id))

        zf.writestr(
            "registers/risk_register.xlsx",
            _build_excel("Risk Register", _build_risk_sheet, risks),
        )
        zf.writestr(
            "registers/asset_register.xlsx",
            _build_excel("Asset Register", _build_asset_sheet, assets),
        )
        zf.writestr(
            "registers/supplier_register.xlsx",
            _build_excel("Supplier Register", _build_supplier_sheet, suppliers),
        )
        zf.writestr(
            "registers/data_processing_register.xlsx",
            _build_excel("Data Processing", _build_data_processing_sheet, dp_records),
        )
        if framework == "ISO_42001":
            zf.writestr(
                "registers/ai_system_register.xlsx",
                _build_excel("AI System Register", _build_ai_system_sheet, ai_systems),
            )
        zf.writestr(
            "registers/statement_of_applicability.xlsx",
            _build_excel("SoA", _build_soa_sheet, controls),
        )
        zf.writestr(
            "evidence/evidence_register.xlsx",
            _build_excel("Evidence Register", _build_evidence_sheet, evidence),
        )

    return buf.getvalue()
